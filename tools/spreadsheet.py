import csv
import re
from pathlib import Path
from tools import Tool
from openpyxl import load_workbook, Workbook

def _resolve(gate, p): return gate.sandbox_root.joinpath(p).resolve()

def _infer_col_type(values) -> str:
    """对一列前几行数据做简单类型推断：int/float/str。"""
    kinds = set()
    for v in values:
        s = str(v).strip()
        if not s:
            continue
        if re.fullmatch(r"[+-]?\d+", s):
            kinds.add("int")
        else:
            try:
                float(s)
                kinds.add("float")
            except ValueError:
                kinds.add("str")
    if not kinds or "str" in kinds:
        return "str"
    if kinds == {"int"}:
        return "int"
    return "float"

def make_spreadsheet_tools(gate) -> list[Tool]:
    def sheet_list(args):
        root = _resolve(gate, ".")
        if not gate.read(root):
            return "拒绝"
        files = [str(x.relative_to(gate.sandbox_root)) for x in root.glob("*")
                 if x.suffix.lower() in (".xlsx", ".csv") and not gate._is_secret(x)]
        return "\n".join(sorted(files)) or "（无表格文件）"

    def _read_xlsx(path, sheet=None):
        wb = load_workbook(path, data_only=True)
        ws = wb[sheet] if sheet else wb.active
        return "\n".join("\t".join("" if c is None else str(c) for c in row) for row in ws.iter_rows(values_only=True))

    def _read_csv(path):
        with open(path, newline="", encoding="utf-8-sig") as f:
            return "\n".join("\t".join(r) for r in csv.reader(f))

    def sheet_read(args):
        p = _resolve(gate, args["path"])
        if not gate.read(p): return "拒绝"
        return _read_xlsx(p, args.get("sheet")) if p.suffix.lower() == ".xlsx" else _read_csv(p)

    def sheet_write_cell(args):
        p = _resolve(gate, args["path"])
        if p.exists():
            if not gate.overwrite(p): return "拒绝"
        elif not gate.write(p):
            return "拒绝"
        if not p.exists():
            Workbook().save(p)
        wb = load_workbook(p)
        ws = wb[args.get("sheet", wb.sheetnames[0])] if args.get("sheet") in wb.sheetnames else wb.create_sheet(args.get("sheet", "Sheet1"))
        ws[args["cell"]] = args["value"]
        wb.save(p)
        return f"已写入 {p.name} {args.get('sheet','Sheet1')}!{args['cell']}"

    def sheet_add_row(args):
        p = _resolve(gate, args["path"])
        if p.exists():
            if not gate.overwrite(p): return "拒绝"
        elif not gate.write(p):
            return "拒绝"
        wb = load_workbook(p) if p.exists() else Workbook()
        ws = wb[args.get("sheet", wb.sheetnames[0])] if args.get("sheet") in wb.sheetnames else wb.create_sheet(args.get("sheet", "Sheet1"))
        ws.append(args["row"])
        wb.save(p)
        return f"已追加 {len(args['row'])} 列到 {p.name}"

    def sheet_summary(args):
        p = _resolve(gate, args["path"])
        if not gate.read(p): return "拒绝"
        wb = load_workbook(p, read_only=True)
        lines = [f"工作表: {', '.join(wb.sheetnames)}"]
        for name in wb.sheetnames:
            ws = wb[name]
            rows = list(ws.iter_rows(values_only=True))
            headers = ["" if c is None else str(c) for c in (rows[0] if rows else ())]
            data_rows = rows[1:6]  # 前 5 行数据做类型推断
            inferred = []
            for col_idx in range(len(headers)):
                vals = [r[col_idx] for r in data_rows if col_idx < len(r) and r[col_idx] is not None]
                inferred.append(_infer_col_type(vals))
            cols = ", ".join(f"{h}({t})" for h, t in zip(headers, inferred)) or "（无列）"
            lines.append(f"- {name}: {ws.max_row} 行 x {ws.max_column} 列")
            lines.append(f"  列: {cols}")
        wb.close()
        return "\n".join(lines)

    return [
        Tool("sheet_list", "列出工作目录下的 xlsx/csv 表格",
             {"type": "object", "properties": {}, "required": []}, sheet_list),
        Tool("sheet_read", "读表格内容（制表符分隔）",
             {"type": "object", "properties": {"path": {"type": "string"}, "sheet": {"type": "string"}}, "required": ["path"]}, sheet_read),
        Tool("sheet_write_cell", "写单元格",
             {"type": "object", "properties": {"path": {"type": "string"}, "sheet": {"type": "string"}, "cell": {"type": "string"}, "value": {"type": "string"}}, "required": ["path", "cell", "value"]}, sheet_write_cell),
        Tool("sheet_add_row", "追加一行",
             {"type": "object", "properties": {"path": {"type": "string"}, "sheet": {"type": "string"}, "row": {"type": "array", "items": {"type": "string"}}}, "required": ["path", "row"]}, sheet_add_row),
        Tool("sheet_summary", "表格结构概览（行数/列数/工作表）",
             {"type": "object", "properties": {"path": {"type": "string"}}, "required": ["path"]}, sheet_summary),
    ]
